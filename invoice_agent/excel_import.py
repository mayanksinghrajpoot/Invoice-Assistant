"""
Read invoice line items from an uploaded spreadsheet or JSON.

Expected columns / keys (any case): name / item / product,
price / rate / unit_price, and qty / quantity. Extra fields are ignored.
"""

from __future__ import annotations

import csv
import io
import json
from typing import Any


class SheetError(ValueError):
    """The file is readable but does not look like an item list."""


NAME_KEYS = ("name", "item", "product", "description", "item_name")
PRICE_KEYS = ("price", "rate", "unit_price", "unitprice", "amount")
QTY_KEYS = ("qty", "quantity", "qty.", "units", "count")


def _norm(cell: Any) -> str:
    return str(cell or "").strip().lower().replace(" ", "_")


def _pick(header: list[str], keys: tuple[str, ...]) -> int | None:
    for i, col in enumerate(header):
        if col in keys:
            return i
    return None


def _pick_key(row: dict[str, Any], keys: tuple[str, ...]) -> Any:
    mapped = {_norm(k): v for k, v in row.items()}
    for key in keys:
        if key in mapped and mapped[key] not in (None, ""):
            return mapped[key]
    return None


def parse_rows(header: list[str], rows: list[list[Any]]) -> list[tuple[str, float, int]]:
    name_i = _pick(header, NAME_KEYS)
    price_i = _pick(header, PRICE_KEYS)
    qty_i = _pick(header, QTY_KEYS)
    if name_i is None or price_i is None:
        raise SheetError(
            "Need columns named like name/item, price/rate, and optionally qty. "
            f"Got: {', '.join(h for h in header if h) or '(empty header)'}"
        )
    items: list[tuple[str, float, int]] = []
    for row in rows:
        if not row or all(str(c).strip() == "" for c in row):
            continue
        name = str(row[name_i] if name_i < len(row) else "").strip()
        if not name:
            continue
        try:
            price = float(row[price_i] if price_i < len(row) else 0)
        except (TypeError, ValueError):
            continue
        qty = 1
        if qty_i is not None and qty_i < len(row) and str(row[qty_i]).strip() != "":
            try:
                qty = int(float(row[qty_i]))
            except (TypeError, ValueError):
                qty = 1
        if qty <= 0 or price < 0:
            continue
        items.append((name, price, qty))
    if not items:
        raise SheetError("No usable rows. Each row needs a name and a price.")
    return items


def _item_from_dict(row: dict[str, Any]) -> tuple[str, float, int] | None:
    name = _pick_key(row, NAME_KEYS)
    price = _pick_key(row, PRICE_KEYS)
    if name is None or price is None:
        return None
    name_s = str(name).strip()
    try:
        price_f = float(price)
    except (TypeError, ValueError):
        return None
    qty = _pick_key(row, QTY_KEYS)
    qty_i = 1
    if qty not in (None, ""):
        try:
            qty_i = int(float(qty))
        except (TypeError, ValueError):
            qty_i = 1
    if not name_s or qty_i <= 0 or price_f < 0:
        return None
    return name_s, price_f, qty_i


def from_json_bytes(data: bytes) -> list[tuple[str, float, int]]:
    try:
        payload = json.loads(data.decode("utf-8-sig"))
    except json.JSONDecodeError as exc:
        raise SheetError(f"JSON is not valid: {exc}") from exc

    if isinstance(payload, dict):
        for key in ("items", "products", "rows", "data", "line_items"):
            if isinstance(payload.get(key), list):
                payload = payload[key]
                break
        else:
            payload = [payload]

    if not isinstance(payload, list):
        raise SheetError(
            "JSON should be a list of items, or {\"items\": [ ... ]}. "
            "Each item needs name/item and price/rate."
        )

    items: list[tuple[str, float, int]] = []
    for row in payload:
        if not isinstance(row, dict):
            continue
        parsed = _item_from_dict(row)
        if parsed:
            items.append(parsed)
    if not items:
        raise SheetError(
            "No usable JSON items. Example: "
            '[{"name": "pen", "price": 40, "qty": 2}]'
        )
    return items


def from_csv_bytes(data: bytes) -> list[tuple[str, float, int]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    table = [[c for c in row] for row in reader]
    if not table:
        raise SheetError("CSV file is empty.")
    header = [_norm(c) for c in table[0]]
    return parse_rows(header, table[1:])


def from_xlsx_bytes(data: bytes) -> list[tuple[str, float, int]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    table: list[list[Any]] = []
    for row in ws.iter_rows(values_only=True):
        table.append(list(row))
    wb.close()
    if not table:
        raise SheetError("Excel sheet is empty.")
    header = [_norm(c) for c in table[0]]
    return parse_rows(header, table[1:])


def from_upload(filename: str, data: bytes) -> list[tuple[str, float, int]]:
    lower = filename.lower()
    if lower.endswith(".csv"):
        return from_csv_bytes(data)
    if lower.endswith(".xlsx"):
        return from_xlsx_bytes(data)
    if lower.endswith(".json"):
        return from_json_bytes(data)
    raise SheetError("Use an .xlsx, .csv, or .json file.")
